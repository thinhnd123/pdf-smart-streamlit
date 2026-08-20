import os
import re
import zipfile
import tempfile
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

# ==============================================================================
# HÀM HỖ TRỢ XỬ LÝ CHUỖI & DÒNG EXCEL
# ==============================================================================
def clean_filename(filename):
    """Loại bỏ ký tự không hợp lệ trong tên file Windows"""
    return re.sub(r'[\\/*?:"<>|]', '_', str(filename)).strip()

def clean_slash(value):
    """Nếu giá trị là '/' (hoặc có chứa khoảng trắng xung quanh '/') thì đổi thành 'NA'"""
    if pd.isna(value) or value is None:
        return ""
    val_str = str(value).strip()
    # Kiểm tra nếu sau khi xóa khoảng trắng chỉ còn đúng 1 ký tự '/'
    if val_str == "/":
        return "NA"
    return val_str

def format_date_mmm(value):
    """Chuyển đổi ngày tháng về định dạng DD/MMM/YYYY (ví dụ: 15/Jan/2026)"""
    if pd.isna(value) or value is None or str(value).strip() == "/":
        return ""
    dt = pd.to_datetime(value, errors='coerce')
    if pd.isna(dt):
        return str(value).strip()
    return dt.strftime("%d/%b/%Y")

def set_wrap_text(cell):
    """Bật tính năng xuống dòng tự động (Wrap Text) giữ nguyên căn lề cũ"""
    current_alignment = cell.alignment
    cell.alignment = Alignment(
        horizontal=current_alignment.horizontal,
        vertical=current_alignment.vertical,
        wrap_text=True
    )

def adjust_row_height(ws, row_idx, texts, char_limit_per_line=25, base_height=20, line_height=15):
    """Tự động tính toán và chỉnh chiều cao dòng chuẩn xác khi chữ dài HOẶC có chứa dấu \n"""
    max_lines = 1
    for txt in texts:
        if txt:
            paragraphs = str(txt).split('\n')
            total_lines_for_cell = 0
            for p in paragraphs:
                p_lines = (len(p) // char_limit_per_line) + 1
                total_lines_for_cell += p_lines
            if total_lines_for_cell > max_lines:
                max_lines = total_lines_for_cell
    
    if max_lines > 1:
        new_height = base_height + (max_lines - 1) * line_height
        current_h = ws.row_dimensions[row_idx].height or base_height
        ws.row_dimensions[row_idx].height = max(current_h, new_height)

def get_row_index(cell_name):
    """Lấy chỉ số dòng từ tên ô (ví dụ 'J11' -> 11)"""
    match = re.search(r'\d+', cell_name)
    return int(match.group()) if match else 1

# ==============================================================================
# QUÉT DỮ LIỆU TỪ TỆP PDF
# ==============================================================================
# Danh sách pattern nhận diện mã tài liệu (Cấu hình mở rộng tại đây)
PATTERNS = [
    r'[A-Za-z0-9]+/[A-Za-z0-9\(\)\-]+',        # Dạng cũ: GST/TD(B)-L001
    r'[A-Za-z0-9]+(?:_[A-Za-z0-9]+)+-[0-9]+',   # Dạng mới: GST_DL_E020-2025, GST_DL_L026-2026
    # r'PATTERN_MOI_TRONG_TUONG_LAI',          # Thêm quy tắc mới vào đây khi cần
]

# Gộp các pattern thành một biểu thức Regex duy nhất bằng toán tử OR (|)
COMBINED_PATTERN = re.compile("|".join(f"(?:{p})" for p in PATTERNS))

def extract_code_from_pdf(pdf_file_obj):
    """Đọc PDF từ bộ nhớ/máy tạm và nhặt tất cả mã quy trình tại Mục 4"""
    try:
        with pdfplumber.open(pdf_file_obj) as pdf:
            page = pdf.pages[1] if len(pdf.pages) > 1 else pdf.pages[0]
            text = page.extract_text()
            if not text:
                return ""

            lines = text.split('\n')
            found_section_4 = False
            extracted_codes = []

            for line in lines:
                line_str = line.strip()
                if "4. Tài liệu cơ sở" in line_str or "Reference documents" in line_str:
                    found_section_4 = True
                    continue

                if found_section_4 and (line_str.startswith("5.") or "5. Thiết bị" in line_str or "5. Local" in line_str):
                    break

                if found_section_4 and line_str:
                    # Tìm tất cả các mã khớp với danh sách PATTERNS trên từng dòng
                    matches = COMBINED_PATTERN.findall(line_str)
                    for code in matches:
                        if code not in extracted_codes:
                            extracted_codes.append(code)

            return "\n".join(extracted_codes)
    except Exception as e:
        return ""

# ==============================================================================
# QUÉT DỮ LIỆU TỪ TỆP PDF (ĐÃ BỔ SUNG GHI LOG CHUẨN CODE GỐC)
# ==============================================================================
def scan_pdf_files(pdf_files):
    """Tạo Dictionary map: {Mã_GCN: Mã_Tài_Liệu} và danh sách Log quét PDF"""
    pdf_map = {}
    pdf_logs = []
    
    total_files = len(pdf_files)
    pdf_logs.append(f"--- ĐANG QUÉT {total_files} FILE PDF TRONG THƯ MỤC ---")
    
    for idx, file in enumerate(pdf_files, 1):
        ma_gcn = os.path.splitext(file.name)[0].strip()
        ma_tai_lieu = extract_code_from_pdf(file)
        pdf_map[ma_gcn] = ma_tai_lieu
        
        # Định dạng dòng log giống code Python gốc
        ma_tl_display = ma_tai_lieu.replace('\n', ' | ') if ma_tai_lieu else ""
        pdf_logs.append(f"[{idx}/{total_files}] PDF: {ma_gcn} -> Mã TL (I20): [{ma_tl_display}]")
        
    return pdf_map, pdf_logs

# ==============================================================================
# HÀM BỎ BIẾN ĐỔI CHUỖI THEO LOẠI (TRANSFORMATION ENGINE)
# ==============================================================================
def apply_transformation(val, transform_type, raw_id="", pdf_data_map=None):
    """
    Áp dụng logic biến đổi dữ liệu tùy theo Kiểu (Transform Type) được chọn trên UI
    """
    # 1. Làm sạch giá trị ban đầu (chuyển '/' thành 'NA')
    str_val = clean_slash(val)

    # 2. Xử lý các logic đặc biệt
    if transform_type == "Cắt lấy phần sau dấu '/'":
        if str_val == "NA":
            return "NA"
        return str_val.split("/")[-1].strip() if "/" in str_val else str_val

    elif transform_type == "Tạo mã 'M' (Prefix M)":
        parts = raw_id.split('.')
        return f"M{parts[1]}" if len(parts) >= 2 else "M"

    elif transform_type == "Định dạng Ngày (DD/MMM/YYYY)":
        return format_date_mmm(val)

    elif transform_type == "Tra cứu PDF theo Mã GCN":
        return pdf_data_map.get(str_val, "") if pdf_data_map else ""

    elif transform_type == "Viết HOA toàn bộ":
        return str_val.upper()

    elif transform_type == "Viết thường toàn bộ":
        return str_val.lower()

    elif transform_type == "Đánh tích nhóm '6' (ü)":
        parts = str(raw_id).split('.')
        if len(parts) >= 3 and parts[2].strip().startswith('6'):
            return "ü"
        return ""

    # Mặc định (Nguyên bản Direct): Trả về giá trị đã làm sạch ('/' -> 'NA')
    return str_val

# ==============================================================================
# ENGINE XỬ LÝ CHÍNH
# ==============================================================================
def run_generate_forms(file_tong_bytes, file_form_bytes, pdf_files, mapping_config):
    """
    Xử lý tạo các file Form Excel và đóng gói vào file ZIP (Log chuẩn 100% định dạng cũ)
    """
    try:
        log_lines = []

        # 1. Quét PDF trước nếu có & Lấy danh sách log PDF
        pdf_data_map = {}
        if pdf_files:
            pdf_data_map, pdf_scan_logs = scan_pdf_files(pdf_files)
            log_lines.extend(pdf_scan_logs)
            log_lines.append("") # Dòng trống phân cách
        else:
            log_lines.append("--- KHÔNG CÓ FILE PDF NÀO ĐƯỢC TẢI LÊN ---\n")

        # 2. Đọc file Tổng
        df = pd.read_excel(file_tong_bytes)
        df = df.replace(r'^\s*/\s*$', 'NA', regex=True)
        
        # Đếm tổng số dòng có dữ liệu hợp lệ để hiển thị [stt/tổng]
        id_col = mapping_config.get("id_column")
        valid_rows = [r for _, r in df.iterrows() if id_col in r and pd.notna(r[id_col]) and str(r[id_col]).strip() and str(r[id_col]).strip().lower() != 'nan']
        total_excel_files = len(valid_rows)

        log_lines.append("--- ĐANG ĐỌC FILE TỔNG EXCEL VÀ ĐIỀN DỮ LIỆU ---")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = tempfile.mkdtemp()
        generated_files = []

        # 3. Lặp từng dòng dữ liệu trong File Tổng
        file_count = 0
        for idx, row in df.iterrows():
            if id_col not in row or pd.isna(row[id_col]):
                continue
            
            raw_id = str(row[id_col]).strip()
            if not raw_id or raw_id.lower() == 'nan':
                continue

            file_count += 1

            # Mở workbook form mẫu
            file_form_bytes.seek(0)
            wb = load_workbook(file_form_bytes)
            ws = wb.active

            # Biến lưu giá trị kết quả tra cứu được ở ô I20 (Mã TL) để đưa vào Log
            i20_val_log = ""

            # --- A. Xử lý danh sách Dynamic Mapping Pairs ---
            for pair in mapping_config.get("dynamic_pairs", []):
                col_name = pair.get("excel_col")
                target_cells_str = pair.get("target_cell", "")
                target_cells = [c.strip().upper() for c in target_cells_str.split(",") if c.strip()]
                transform_type = pair.get("transform_type", "Nguyên bản (Direct)")

                if col_name in row and target_cells:
                    raw_val = row[col_name]
                    final_val = apply_transformation(raw_val, transform_type, raw_id=raw_id, pdf_data_map=pdf_data_map)

                    # Bắt giá trị tra cứu để hiển thị đúng cột I20 trong log
                    if transform_type == "Tra cứu PDF theo Mã GCN" or "I20" in target_cells:
                        i20_val_log = final_val.replace('\n', ' | ')

                    for cell_name in target_cells:
                        ws[cell_name] = final_val
                        set_wrap_text(ws[cell_name])
                        adjust_row_height(ws, row_idx=get_row_index(cell_name), texts=[final_val])

            # --- B. Xử lý Rule Cố định Đặc biệt (Đánh tích ü - Chạy sau cùng) ---
            rules = mapping_config.get("special_rules", {})
            if rules.get("checkmark_logic", {}).get("active"):
                cell_k = rules["checkmark_logic"].get("cell_option1", "K14")
                cell_n = rules["checkmark_logic"].get("cell_option2", "N14")
                
                parts = raw_id.split('.')
                ws[cell_k] = ""
                ws[cell_n] = ""
                
                if len(parts) >= 3 and parts[2].strip().startswith('6'):
                    ws[cell_k] = "ü"
                else:
                    ws[cell_n] = "ü"

            # Lưu file Excel theo Mã QL
            safe_filename = clean_filename(raw_id)
            out_file_path = os.path.join(temp_dir, f"{safe_filename}.xlsx")
            wb.save(out_file_path)
            wb.close()
            generated_files.append(out_file_path)

            # Ghi dòng Log theo đúng định dạng chuẩn code cũ:
            # [1/143] Hoàn tất: ACS1.VMQ.501.xlsx (I20: '')
            log_lines.append(f"[{file_count}/{total_excel_files}] Hoàn tất: {safe_filename}.xlsx (I20: '{i20_val_log}')")

        if not generated_files:
            return None, "Không tạo được file nào. Vui lòng kiểm tra lại cột định danh Mã quản lý."

        # Tạo file Log_Doi_Chieu_PDF.txt trong thư mục tạm
        log_file_path = os.path.join(temp_dir, "Log_Doi_Chieu_PDF.txt")
        with open(log_file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(log_lines))

        # 4. Nén tất cả các file Excel + File Log vào ZIP
        zip_path = os.path.join(tempfile.gettempdir(), f"Danh_Sach_Form_Hoan_Thanh_{timestamp}.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in generated_files:
                zipf.write(file_path, arcname=os.path.basename(file_path))
            zipf.write(log_file_path, arcname="Log_Doi_Chieu_PDF.txt")

        # Dọn dẹp temp
        for f in generated_files + [log_file_path]:
            try: os.remove(f)
            except Exception: pass

        return zip_path, f"Đã xuất thành công {len(generated_files)} file Form Excel và 1 file Log!"

    except Exception as e:
        return None, f"Lỗi xử lý: {str(e)}"