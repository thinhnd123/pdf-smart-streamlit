import pandas as pd
import tempfile
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def run_excel_compare(
    file_a,
    file_b,
    keys_a,
    keys_b
):

    try:

        df_a = pd.read_excel(
            file_a,
            dtype=str
        ).fillna("")

        df_b = pd.read_excel(
            file_b,
            dtype=str
        ).fillna("")

        if len(keys_a) != len(keys_b):

            return (
                None,
                None,
                None,
                None,
                None,
                None,
                "Số lượng cột khóa không khớp"
            )

        # ==================================
        # KEY
        # ==================================

        df_a["_KEY"] = (
            df_a[keys_a]
            .astype(str)
            .agg(
                "|".join,
                axis=1
            )
        )

        df_b["_KEY"] = (
            df_b[keys_b]
            .astype(str)
            .agg(
                "|".join,
                axis=1
            )
        )

        set_a = set(
            df_a["_KEY"]
        )

        set_b = set(
            df_b["_KEY"]
        )

        matched = set_a & set_b

        only_a = set_a - set_b

        only_b = set_b - set_a

        # ==================================
        # DATAFRAME KẾT QUẢ
        # ==================================

        df_match = pd.DataFrame(
            sorted(list(matched)),
            columns=["Khóa"]
        )

        df_only_a = pd.DataFrame(
            sorted(list(only_a)),
            columns=["Khóa"]
        )

        df_only_b = pd.DataFrame(
            sorted(list(only_b)),
            columns=["Khóa"]
        )

        # ==================================
        # SO SÁNH CHI TIẾT
        # ==================================

        detail_rows = []

        common_columns = [

            c

            for c in df_a.columns

            if c in df_b.columns

            and c != "_KEY"
        ]

        changed_keys = set()

        for key in matched:

            row_a = (
                df_a[
                    df_a["_KEY"] == key
                ]
                .iloc[0]
            )

            row_b = (
                df_b[
                    df_b["_KEY"] == key
                ]
                .iloc[0]
            )

            for col in common_columns:

                value_a = str(
                    row_a[col]
                ).strip()

                value_b = str(
                    row_b[col]
                ).strip()

                if value_a != value_b:

                    changed_keys.add(
                        key
                    )

                    detail_rows.append(
                        {
                            "Key": key,
                            "Column": col,
                            "File_A": value_a,
                            "File_B": value_b
                        }
                    )

        df_detail = pd.DataFrame(
            detail_rows
        )

        # ==================================
        # AUDIT REPORT
        # ==================================

        if len(df_detail):

            audit_report = (

                df_detail

                .groupby("Column")

                .size()

                .reset_index(
                    name="Changed_Count"
                )

                .sort_values(
                    "Changed_Count",
                    ascending=False
                )
            )

        else:

            audit_report = pd.DataFrame(
                columns=[
                    "Column",
                    "Changed_Count"
                ]
            )

        # ==================================
        # EXPORT
        # ==================================

        output_file = os.path.join(
            tempfile.gettempdir(),
            "Excel_Compare.xlsx"
        )

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            df_match.to_excel(
                writer,
                sheet_name="Matched",
                index=False
            )

            df_only_a.to_excel(
                writer,
                sheet_name="Only_A",
                index=False
            )

            df_only_b.to_excel(
                writer,
                sheet_name="Only_B",
                index=False
            )

            df_detail.to_excel(
                writer,
                sheet_name="Differences",
                index=False
            )

            audit_report.to_excel(
                writer,
                sheet_name="Audit_Report",
                index=False
            )

        # ==================================
        # TÔ MÀU
        # ==================================

        wb = load_workbook(
            output_file
        )

        red_fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE"
        )

        ws = wb["Differences"]

        for row in ws.iter_rows(
            min_row=2
        ):

            for cell in row:

                cell.fill = red_fill

        wb.save(
            output_file
        )

        summary = {

            "match": len(df_match),

            "only_a": len(df_only_a),

            "only_b": len(df_only_b),

            "differences": len(df_detail),

            "changed_keys": len(
                changed_keys
            )
        }

        return (
            output_file,
            df_match,
            df_only_a,
            df_only_b,
            df_detail,
            summary,
            "✅ Audit hoàn tất"
        )

    except Exception as e:

        return (
            None,
            None,
            None,
            None,
            None,
            None,
            str(e)
        )